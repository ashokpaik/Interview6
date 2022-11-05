import openpyxl


class Mainpagedata:
    mainpagedata1 = [{"Name": "Ashok", "Email": "ashokpaik@gmail.com", "Password": "Ashok123", "Birthday": "21-07-1978"},
                    {"Name": "Ashok1", "Email": "ashok1paik@gmail.com", "Password": "Ashok1123",
                     "Birthday": "27-07-1978"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Admin\\PycharmProjects\\Interview6\\testdata\\mainpagedata.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]




